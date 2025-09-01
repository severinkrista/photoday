# statistic.py
# Функции для сбора и отображения статистики

import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk
import state  # Для доступа к пути Excel-файла из настроек
import settings # Для сохранения настроек

# Предполагаемые индексы колонок в Excel (с 1, как в Excel)
DATE_COL_INDEX = 1      # Колонка "Дата"
TASK_TYPE_COL_INDEX = 5 # Колонка "Вид задачи"
DIFFICULTY_COL_INDEX = 7 # Колонка "Сложность"

# Значение по умолчанию для размера статистики
DEFAULT_STATISTIC_SIZE = 5

# Кэширование распарсенных дат для ускорения
_date_cache = {}

def _parse_date_cached(date_value):
    """Парсит дату с кэшированием для ускорения."""
    if date_value in _date_cache:
        return _date_cache[date_value]
    
    record_date = None
    if isinstance(date_value, datetime):
        record_date = date_value.date()
    elif isinstance(date_value, str):
        try:
            # Пытаемся распарсить дату в формате dd.mm.yyyy
            record_date = datetime.strptime(date_value, "%d.%m.%Y").date()
        except ValueError:
            pass
    
    _date_cache[date_value] = record_date
    return record_date

def get_task_statistics():
    """
    Считает статистику по записям из Excel-файла.
    Возвращает только дни, за которые есть хотя бы одна запись.
    
    Возвращает словарь с ключами:
    - 'days_data': dict, где ключ - дата (datetime.date), значение - dict со статистикой по этой дате
                   {'count': int, 'total_difficulty': int, 'difficulty_by_type': dict}
    - 'error': str or None (если ошибка произошла)
    """
    stats = {
        'days_data': {},  # Словарь для хранения данных по дням с записями
        'error': None
    }

    # Получаем путь к Excel-файлу из настроек
    xlsx_path = state.settings.get("excel_path", None)
    if not xlsx_path:
        stats['error'] = "Путь к Excel-файлу не задан в настройках."
        return stats

    xlsx_path_value = xlsx_path.get() if hasattr(xlsx_path, 'get') else xlsx_path
    if not xlsx_path_value:
        stats['error'] = "Путь к Excel-файлу пуст."
        return stats

    if not os.path.exists(xlsx_path_value):
        stats['error'] = f"Excel-файл не найден: {xlsx_path_value}"
        return stats

    # Очистка кэша перед каждым вызовом
    global _date_cache
    _date_cache = {}

    try:
        # Открываем файл в режиме только для чтения и без вычисления формул
        wb = load_workbook(xlsx_path_value, read_only=True, data_only=True)
        ws = wb.active

        # Пропускаем заголовок, если он есть
        start_row = 1
        if ws.max_row > 0:
            first_date_cell = ws.cell(row=1, column=DATE_COL_INDEX).value
            if isinstance(first_date_cell, str) and 'дата' in first_date_cell.lower():
                start_row = 2

        # Используем iter_rows для более эффективного чтения
        for row in ws.iter_rows(min_row=start_row, values_only=True, max_col=DIFFICULTY_COL_INDEX):
            # Получаем дату из кортежа
            date_cell_value = row[DATE_COL_INDEX - 1] if len(row) >= DATE_COL_INDEX else None
            if not date_cell_value:
                continue

            # Используем кэшированную функцию парсинга даты
            record_date = _parse_date_cached(date_cell_value)

            # Если дата не распознана, пропускаем
            if record_date is None:
                continue

            # Получаем вид задачи и сложность из кортежа
            task_type = row[TASK_TYPE_COL_INDEX - 1] if len(row) >= TASK_TYPE_COL_INDEX else None
            if not task_type:
                task_type = "Не указан"

            difficulty_raw = row[DIFFICULTY_COL_INDEX - 1] if len(row) >= DIFFICULTY_COL_INDEX else None
            try:
                difficulty = int(difficulty_raw) if difficulty_raw is not None else 0
            except (ValueError, TypeError):
                difficulty = 0

            # Инициализируем структуру для этой даты, если она еще не существует
            if record_date not in stats['days_data']:
                stats['days_data'][record_date] = {
                    'count': 0,
                    'total_difficulty': 0,
                    'difficulty_by_type': {}
                }

            # Обновляем статистику для этой даты
            stats['days_data'][record_date]['count'] += 1
            stats['days_data'][record_date]['total_difficulty'] += difficulty
            
            if task_type not in stats['days_data'][record_date]['difficulty_by_type']:
                stats['days_data'][record_date]['difficulty_by_type'][task_type] = 0
            stats['days_data'][record_date]['difficulty_by_type'][task_type] += difficulty

    except Exception as e:
        stats['error'] = f"Ошибка при чтении Excel-файла: {e}"
    finally:
        if 'wb' in locals():
            wb.close()

    return stats

def show_statistics(parent_window):
    """
    Собирает и отображает статистику во всплывающем окне в виде таблицы.
    Отображаются только дни, за которые есть записи.
    parent_window: ссылка на главное окно приложения (root), 
                   необходима для создания Toplevel.
    """
    # 1. Вызов функции сбора статистики
    stats_result = get_task_statistics()

    # 2. Создание нового окна для отображения
    stats_window = tk.Toplevel(parent_window)
    stats_window.title("Статистика")
    stats_window.geometry("600x350")
    stats_window.resizable(True, True)
    stats_window.grab_set() # Делает окно модальным
    stats_window.focus_set()

    # 3. === ИЗМЕНЕНО: Получаем statistic_size из settings.ini или используем значение по умолчанию ===
    # Загружаем настройки из settings.ini
    settings_path = settings.get_settings_path()
    statistic_size = DEFAULT_STATISTIC_SIZE  # Значение по умолчанию
    
    if os.path.exists(settings_path):
        import configparser
        config = configparser.ConfigParser()
        config.optionxform = str
        try:
            config.read(settings_path, encoding='utf-8')
            if 'Settings' in config:
                section = config['Settings']
                if 'statistic_size' in section:
                    try:
                        statistic_size = int(section['statistic_size'])
                        # Проверка на допустимый диапазон
                        if statistic_size < 1:
                            statistic_size = 1
                        elif statistic_size > 30:
                            statistic_size = 30
                    except ValueError:
                        statistic_size = DEFAULT_STATISTIC_SIZE
        except Exception as e:
            print(f"Ошибка при чтении настроек для статистики: {e}")
    
    # Переменная для хранения количества дней для отображения
    statistic_size_var = tk.IntVar(value=statistic_size)

    # 4. Создание фрейма для элементов управления количеством дней
    control_frame = tk.Frame(stats_window)
    control_frame.pack(fill="x", padx=10, pady=(10, 5))
    
    tk.Label(control_frame, text="Показать дней:").pack(side="left")
    statistic_size_entry = tk.Entry(control_frame, textvariable=statistic_size_var, width=5)
    statistic_size_entry.pack(side="left", padx=(5, 5))
    
    # Функция для обновления таблицы с новым количеством дней
    def update_table():
        # Получаем новое количество дней
        try:
            current_statistic_size = statistic_size_var.get()
            if current_statistic_size < 1:
                current_statistic_size = 1
                statistic_size_var.set(1)
            elif current_statistic_size > 30:  # Ограничение на случай очень большого числа
                current_statistic_size = 30
                statistic_size_var.set(30)
        except tk.TclError:
            current_statistic_size = DEFAULT_STATISTIC_SIZE
            statistic_size_var.set(DEFAULT_STATISTIC_SIZE)
            
        # === ИЗМЕНЕНО: Сохраняем statistic_size в settings.ini ===
        settings_path = settings.get_settings_path()
        if os.path.exists(settings_path):
            import configparser
            config = configparser.ConfigParser()
            config.optionxform = str
            try:
                config.read(settings_path, encoding='utf-8')
                if 'Settings' not in config:
                    config['Settings'] = {}
                config['Settings']['statistic_size'] = str(current_statistic_size)
                with open(settings_path, 'w', encoding='utf-8') as configfile:
                    config.write(configfile)
                print(f"Настройка statistic_size сохранена в {settings_path}")
            except Exception as e:
                print(f"Ошибка при сохранении настройки statistic_size: {e}")
        # === /ИЗМЕНЕНО ===
            
        # Обновляем отображение данных
        if stats_result['error']:
            # Очищаем таблицу
            for item in tree.get_children():
                tree.delete(item)
            tree.insert('', tk.END, values=('Ошибка получения статистики:',))
        else:
            days_with_data = sorted(stats_result['days_data'].keys(), reverse=True)
            days_to_show_list = days_with_data[:current_statistic_size]
            
            # Очищаем таблицу
            for item in tree.get_children():
                tree.delete(item)
                
            if not days_to_show_list:
                tree.insert('', tk.END, values=('Нет данных для отображения',))
            else:
                # Подготавливаем данные для отображения
                days_data = stats_result['days_data']
                
                # Собираем все уникальные типы задач из отображаемых дней
                all_task_types = set()
                for day in days_to_show_list:
                    all_task_types.update(days_data[day]['difficulty_by_type'].keys())
                
                # Обновляем заголовки колонок
                date_columns = [date.strftime("%d.%m.%Y") for date in days_to_show_list]
                columns = ('Показатель',) + tuple(date_columns)
                
                # Обновляем столбцы в treeview
                tree.configure(columns=columns)
                tree.heading('Показатель', text='Показатель')
                for date_str in date_columns:
                    tree.heading(date_str, text=date_str)
                
                # Настройка ширин колонок
                tree.column('Показатель', width=150, anchor='w')
                for date_str in date_columns:
                    tree.column(date_str, width=70, anchor='center') # Увеличена ширина для лучшей читаемости
                
                # Добавляем строки в таблицу
                # Всего записей
                row_values = ['Всего записей:']
                for day in days_to_show_list:
                    row_values.append(days_data[day]['count'])
                tree.insert('', tk.END, values=tuple(row_values))
                
                # Сумма сложностей
                row_values = ['Сумма сложностей:']
                for day in days_to_show_list:
                    row_values.append(days_data[day]['total_difficulty'])
                tree.insert('', tk.END, values=tuple(row_values))
                
                # Сложность по типам
                if all_task_types:
                     tree.insert('', tk.END, values=('',) * (len(days_to_show_list) + 1)) # Пустая строка-разделитель
                     tree.insert('', tk.END, values=('Сложность по типам:',) + ('',) * len(days_to_show_list))
                     for task_type in sorted(all_task_types): # Сортируем для порядка
                        row_values = [f"  - {task_type}"]
                        for day in days_to_show_list:
                            difficulty = days_data[day]['difficulty_by_type'].get(task_type, 0)
                            row_values.append(difficulty)
                        tree.insert('', tk.END, values=tuple(row_values))
                else:
                     tree.insert('', tk.END, values=('Сложность по типам:',) + ('Нет данных',) * len(days_to_show_list))

    # Кнопка "Применить" с зеленой галочкой
    apply_button = tk.Button(control_frame, text=" ✓ ", command=update_table, bg="#4CAF50", fg="white")
    apply_button.pack(side="left", padx=(5, 0))

    # 5. Создание виджета Treeview для таблицы
    if stats_result['error']:
        # Если ошибка, создаем таблицу с одной колонкой для отображения сообщения
        columns = ('Показатель',)
        tree = ttk.Treeview(stats_window, columns=columns, show='headings', height=5)
        tree.heading('Показатель', text='Показатель')
        tree.column('Показатель', width=550, anchor='w')
    else:
        days_with_data = sorted(stats_result['days_data'].keys(), reverse=True)
        # === ИЗМЕНЕНО: Используем statistic_size_var.get() вместо days_to_show_default ===
        days_to_show_list = days_with_data[:statistic_size_var.get()]
        
        if not days_to_show_list:
            # Если нет дней с данными, создаем таблицу с одной колонкой
            columns = ('Показатель',)
            tree = ttk.Treeview(stats_window, columns=columns, show='headings', height=5)
            tree.heading('Показатель', text='Показатель')
            tree.column('Показатель', width=550, anchor='w')
        else:
            # Форматируем даты в строку dd.mm.yyyy для заголовков
            date_columns = [date.strftime("%d.%m.%Y") for date in days_to_show_list]
            columns = ('Показатель',) + tuple(date_columns)
            
            tree = ttk.Treeview(stats_window, columns=columns, show='headings', height=12) # Уменьшена высота
            
            # Определение заголовков
            tree.heading('Показатель', text='Показатель')
            for date_str in date_columns:
                tree.heading(date_str, text=date_str)
            
            # Настройка ширин колонок
            tree.column('Показатель', width=150, anchor='w')
            for date_str in date_columns:
                tree.column(date_str, width=70, anchor='center') # Увеличена ширина

    # Добавление скроллбара
    scrollbar = ttk.Scrollbar(stats_window, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscroll=scrollbar.set)

    # 6. Заполнение таблицы данными (начальное отображение)
    if stats_result['error']:
        tree.insert('', tk.END, values=('Ошибка получения статистики:',))
    else:
        # === ИЗМЕНЕНО: Используем statistic_size_var.get() для начального отображения ===
        days_with_data = sorted(stats_result['days_data'].keys(), reverse=True)
        days_to_show_list = days_with_data[:statistic_size_var.get()]
        
        if not days_to_show_list:
            tree.insert('', tk.END, values=('Нет данных для отображения',))
        else:
            # Подготавливаем данные для отображения
            days_data = stats_result['days_data']
            
            # Собираем все уникальные типы задач из отображаемых дней
            all_task_types = set()
            for day in days_to_show_list:
                all_task_types.update(days_data[day]['difficulty_by_type'].keys())
            
            # Добавляем строки в таблицу
            # Всего записей
            row_values = ['Всего записей:']
            for day in days_to_show_list:
                row_values.append(days_data[day]['count'])
            tree.insert('', tk.END, values=tuple(row_values))
            
            # Сумма сложностей
            row_values = ['Сумма сложностей:']
            for day in days_to_show_list:
                row_values.append(days_data[day]['total_difficulty'])
            tree.insert('', tk.END, values=tuple(row_values))
            
            # Сложность по типам
            if all_task_types:
                 tree.insert('', tk.END, values=('',) * (len(days_to_show_list) + 1)) # Пустая строка-разделитель
                 tree.insert('', tk.END, values=('Сложность по типам:',) + ('',) * len(days_to_show_list))
                 for task_type in sorted(all_task_types): # Сортируем для порядка
                    row_values = [f"  - {task_type}"]
                    for day in days_to_show_list:
                        difficulty = days_data[day]['difficulty_by_type'].get(task_type, 0)
                        row_values.append(difficulty)
                    tree.insert('', tk.END, values=tuple(row_values))
            else:
                 tree.insert('', tk.END, values=('Сложность по типам:',) + ('Нет данных',) * len(days_to_show_list))

    # 7. Размещение виджетов в окне
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=(0, 10))
