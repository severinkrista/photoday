# file_operations.py
# Операции с файлами: открытие, сохранение

import os
import subprocess
import platform
import tkinter.messagebox as messagebox # Импортируем messagebox
import state # Для доступа к путям настроек

# Импортируем openpyxl внутри функций, которые его используют, чтобы избежать импорта, если не используется
# from openpyxl import load_workbook, Workbook
# from openpyxl.styles import Alignment

def open_text():
    """Открывает текстовый файл с использованием системного приложения."""
    path = state.settings["txt_path"].get()
    if not path:
        messagebox.showwarning("Ошибка", "Не указан путь к текстовому файлу.")
        return
    if not os.path.exists(path):
        messagebox.showwarning("Файл не найден", f"Файл не существует:\n{path}")
        return
    try:
        if platform.system() == "Windows":
            os.startfile(path)
        elif platform.system() == "Darwin":
            subprocess.run(["open", path])
        else:
            subprocess.run(["xdg-open", path])
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось открыть файл:\n{e}")

def open_excel():
    """Открывает Excel файл с использованием системного приложения."""
    path = state.settings["excel_path"].get()
    if not path:
        messagebox.showwarning("Ошибка", "Не указан путь к Excel файлу.")
        return
    if not os.path.exists(path):
        messagebox.showwarning("Файл не найден", f"Файл не существует:\n{path}")
        return
    try:
        if platform.system() == "Windows":
            os.startfile(path)
        elif platform.system() == "Darwin":
            subprocess.run(["open", path])
        else:
            subprocess.run(["xdg-open", path])
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось открыть файл:\n{e}")

def save_records_to_txt(record_widgets):
    """Сохраняет записи в текстовый файл."""
    path = state.settings["txt_path"].get().strip()
    if not path:
        messagebox.showwarning("Ошибка", "Не указан путь для TXT-файла.")
        return False

    try:
        # Проверка и создание директории
        txt_dir = os.path.dirname(path)
        if txt_dir and not os.path.exists(txt_dir):
            try:
                os.makedirs(txt_dir)
            except Exception as e:
                 messagebox.showwarning("Предупреждение", f"Не удалось создать директорию для TXT файла:\n{txt_dir}\nОшибка: {e}")
                 return False # Не продолжаем сохранение, если не удалось создать директорию

        with open(path, "a", encoding="utf-8") as f:
            for rec in record_widgets:
                desc = rec['description_text'].get("1.0", "end-1c").strip()
                if not desc:
                    continue
                # === ИЗМЕНЕНО: Убираем переносы строк из описания ===
                desc_single_line = desc.replace('\n', ' ').replace('\r', ' ')
                line = f"{rec['date_var'].get()}\t{rec['time_var'].get()}\t{rec['weekday_var'].get()}\t" \
                       f"{rec['part_of_day_var'].get()}\t{rec['task_type_var'].get()}\t{desc_single_line}\t{rec['difficulty_var'].get()}"
                f.write(line + "\n")
        return True
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить в TXT:\n{e}")
        return False

def save_records_to_excel(record_widgets):
    """Сохраняет записи в Excel файл."""
    path = state.settings["excel_path"].get().strip()
    if not path:
        messagebox.showwarning("Ошибка", "Не указан путь для Excel-файла.")
        return False

    try:
        # Проверка и создание директории
        xlsx_dir = os.path.dirname(path)
        if xlsx_dir and not os.path.exists(xlsx_dir):
            try:
                os.makedirs(xlsx_dir)
            except Exception as e:
                 messagebox.showwarning("Предупреждение", f"Не удалось создать директорию для XLSX файла:\n{xlsx_dir}\nОшибка: {e}")
                 return False # Не продолжаем сохранение, если не удалось создать директорию

        from openpyxl import load_workbook, Workbook # Импортируем здесь
        
        headers = ["Дата", "Время", "День недели", "Часть дня", "Вид задачи", "Задача", "Сложность"]
        wb = load_workbook(path) if os.path.exists(path) else Workbook()
        ws = wb.active
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            ws.append(headers)
        for rec in record_widgets:
            desc = rec['description_text'].get("1.0", "end-1c").strip()
            if not desc:
                continue
            # === ИЗМЕНЕНО: Преобразуем сложность в число ===
            try:
                difficulty_value = int(rec['difficulty_var'].get())
            except ValueError:
                # Если не удалось преобразовать, сохраняем как есть (строку)
                difficulty_value = rec['difficulty_var'].get()
            # === ИЗМЕНЕНО: Убираем переносы строк из описания ===
            desc_single_line = desc.replace('\n', ' ').replace('\r', ' ')
            ws.append([
                rec['date_var'].get(),
                rec['time_var'].get(),
                rec['weekday_var'].get(),
                rec['part_of_day_var'].get(),
                rec['task_type_var'].get(),
                desc_single_line,  # Используем обработанное описание
                difficulty_value  # Используем числовое значение
            ])
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        wb.save(path)
        return True
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить в Excel:\n{e}")
        return False
